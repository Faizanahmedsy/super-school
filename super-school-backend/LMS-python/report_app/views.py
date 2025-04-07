from django.http import HttpResponse
from nest_db_app.models import *
from nest_db_app.decorators import permission_required
from rest_framework.decorators import action
from rest_framework import viewsets
from openpyxl import Workbook
from django.db.models import Count, Q ,Sum ,Avg,Prefetch
from django.utils.timezone import is_aware
from datetime import datetime
from exam_app.models import *
from personalized_learning.models import *
# Create your views here.
from django.conf import settings
from rest_framework.response import Response
from django_filters.rest_framework import DjangoFilterBackend,OrderingFilter
from exam_app.pagination import GlobalPagination
from rest_framework.filters import SearchFilter
from django.http import FileResponse
import io
from django.utils.translation import gettext as _

class ReportViewset(viewsets.GenericViewSet):
    pagination_class = GlobalPagination
    download_query_param = "download"

    @action(detail=False, methods=["get"], url_path="generate-school-report")
    @permission_required('reports_management', 'view')
    def generate_school_report(self, request, *args, **kwargs):
        if request.user.role.role_name not in settings.SUPER_ROLE:
            return Response({"error": _("You do not have permission to generate reports.")}, status=403)
        # Base QuerySet for School
        queryset = School.objects.all()

        # Apply Filters
        school_name = request.query_params.get("school_name")
        if school_name:
            queryset = queryset.filter(school_name__icontains=school_name)

        school_id = request.query_params.get("school_id")
        if school_id:
            queryset = queryset.filter(id=school_id)

        created_at_gte = request.query_params.get("created_at__gte")
        if created_at_gte:
            queryset = queryset.filter(created_at__gte=created_at_gte)

        created_at_lte = request.query_params.get("created_at__lte")
        if created_at_lte:
            queryset = queryset.filter(created_at__lte=created_at_lte)

        # Filter by Year (created_at year)
        year = request.query_params.get("year")
        if year:
            try:
                year = int(year)
                queryset = queryset.filter(id__in=School.objects.filter(created_at__year=year).values_list('id', flat=True))
            except ValueError:
                return Response({"error": _("Invalid year format. Please provide a valid year.")}, status=400)

        # Filter Schools that have or don't have associated Teachers
        has_teachers = request.query_params.get("has_teachers")
        if has_teachers == "true":
            queryset = queryset.filter(teachers_school__isnull=False).distinct()
        elif has_teachers == "false":
            queryset = queryset.filter(teachers_school__isnull=True).distinct()

        # Filter Schools that have or don't have Events
        has_events = request.query_params.get("has_events")
        if has_events == "true":
            queryset = queryset.filter(calenders_school__isnull=False).distinct()
        elif has_events == "false":
            queryset = queryset.filter(calenders_school__isnull=True).distinct()

        # Apply filtering based on SearchFilter and OrderingFilter
        search = self.request.query_params.get("search")
        if search:
            queryset = queryset.filter(school_name__icontains=search)

        ordering = self.request.query_params.get("ordering", "school_name")  # Default ordering by school_name
        if ordering:
            queryset = queryset.order_by(ordering)

        # Construct the Final Data Payload
        school_data = []
        for school in queryset:
            try:
                # Fetch counts and related fields
                school_admins_count = school.school_admins_school.count()
                teachers_count = school.teachers_school.count()
                learners_count = school.students_school.count()
                parents_count = school.parents_school.count()
                total_subjects = school.subjects_school.count()
                total_assessments = school.school_assessments.count()
                total_assessment_subjects = school.assessment_subjects_school.count()
                total_answersheets_uploaded = school.student_answer_sheets_school.count()
                quizzes_by_type = school.quiz_set.values("quiz_type").annotate(total=Count("quiz_type"))
                total_events = school.calenders_school.count()
                total_lesson_plans = school.lesson_plans_school.count()
                # total_study_materials = school.study_materials_school.count()

                quiz_types = {q.get("quiz_type"): q.get("total", 0) for q in quizzes_by_type}

                # Append school data to the result
                school_data.append({
                    "school_id": school.id,
                    "school_name": school.school_name,
                    "school_admins_count": school_admins_count,
                    "teachers_count": teachers_count,
                    "learners_count": learners_count,
                    "parents_count": parents_count,
                    "total_subjects": total_subjects,
                    "total_assessments": total_assessments,
                    "total_assessment_subjects": total_assessment_subjects,
                    "total_answersheets_uploaded": total_answersheets_uploaded,
                    "quiz_types": quiz_types,
                    "total_events": total_events,
                    "total_lesson_plans": total_lesson_plans,
                    # "total_study_materials": total_study_materials,
                })
            except Exception as e:
                return Response(
                    {"error": _("Error processing data for school %(school_name)s: %(error)s") %
                              {"school_name": school.school_name, "error": str(e)}},
                    status=500
                )

        download = request.query_params.get("download", "false").lower() == "true"
        if download:
            import io
            from django.http import FileResponse
            from openpyxl import Workbook

            output = io.BytesIO()
            wb = Workbook()
            ws = wb.active
            ws.title = "School Report"

            # Extract headers from the first entry of school_data
            headers = list(school_data[0].keys()) if school_data else []
            ws.append(headers)

            for row in school_data:
                excel_row = []

                # Ensure all fields in the row are in a supported format
                for key in headers:
                    value = row.get(key)

                    # Convert dictionaries and unsupported types to strings
                    if isinstance(value, dict):
                        value = str(value)  # Convert dict to string
                    elif isinstance(value, list):
                        value = ", ".join(map(str, value))  # Convert list to comma-separated string
                    excel_row.append(value)

                ws.append(excel_row)  # Add the processed row to the worksheet

            wb.save(output)
            output.seek(0)
            return FileResponse(output, as_attachment=True, filename="school_report.xlsx")

        # Apply Pagination
        paginator = self.pagination_class()
        paginated_data = paginator.paginate_queryset(school_data, request)
        return paginator.get_paginated_response(paginated_data)


    @action(detail=False, methods=["get"], url_path="school-teachers-report")
    @permission_required('reports_management', 'view')
    def export_teachers(self, request, *args, **kwargs):
        # Determine the school_id based on user role
        if request.user.role.role_name in settings.SUPER_ROLE:
            school_id = request.query_params.get("school_id")
            if not school_id:
                return Response({"error": _("School ID is required for users with SUPER_ROLE.")}, status=400)
        else:
            school_id = request.user.school_id

        # Begin the initial queryset
        queryset = ClassSubject.objects.select_related(
            "teacher", "school", "grade", "grade_class", "subject"
        )

        # Filter by school_id
        if school_id:
            queryset = queryset.filter(school_id=school_id)
        else:
            return Response({"error": _("School ID is required.")}, status=400)

        # Apply year filter
        batch_id = request.query_params.get("batch_id")
        if batch_id:
            try:
                batch_id = int(batch_id)
                queryset = queryset.filter(batch_id=batch_id)
            except ValueError:
                return Response({"error": _("Invalid batch_id format. It must be an integer.")}, status=400)

        # Apply additional filters
        school_name = request.query_params.get("school_name")
        if school_name:
            queryset = queryset.filter(school__school_name__icontains=school_name)

        # New Filters for grade, grade_class, and subject_id
        grade = request.query_params.get("grade")
        if grade:
            queryset = queryset.filter(grade=grade)

        grade_class = request.query_params.get("grade_class")
        if grade_class:
            queryset = queryset.filter(grade_class=grade_class)

        subject_id = request.query_params.get("subject_id")
        if subject_id:
            try:
                queryset = queryset.filter(subject=subject_id)
            except ValueError:
                return Response({"error": _("Invalid subject_id format. It must be an integer.")}, status=400)

        # Apply search functionality
        search_query = request.query_params.get("search")
        if search_query:
            queryset = queryset.filter(
                Q(teacher__first_name__icontains=search_query) |
                Q(teacher__last_name__icontains=search_query) |
                Q(teacher__email__icontains=search_query)
            )

        # Apply ordering
        order_by = request.query_params.get("ordering", "teacher__first_name")  # Default ordering by teacher name
        reverse = order_by.startswith("-")  # Check if descending
        if reverse:
            order_by = order_by[1:]  # Remove `-` sign for field name
        if order_by in ["teacher__first_name", "teacher__last_name", "teacher__email", "teacher__id"]:
            queryset = queryset.order_by("-" + order_by if reverse else order_by)

        # Group by unique teacher data
        teacher_data = {}
        for record in queryset:
            if record.teacher is not None:
                teacher = record.teacher
                if teacher.id not in teacher_data:
                    teacher_data[teacher.id] = {
                        "teacher_id": teacher.id,
                        "teacher_name": f"{teacher.first_name} {teacher.last_name}".strip(),
                        "email": teacher.email or "N/A",
                        "mobile": teacher.mobile_number or "N/A",
                        "school_name": record.school.school_name,
                        "subjects": set(),
                        "grades_classes": set(),  # Combine grade and class
                    }

                print("grade class", record.grade_class)
                # Add subjects, grades, and grade classes
                if record.subject and record.subject.master_subject:
                    subject_detail = (
                        f"{record.subject.master_subject.subject_name} ({record.subject.master_subject.subject_code})"
                    )
                    teacher_data[teacher.id]["subjects"].add(subject_detail)

                # Combine grades and classes
                if record.grade and record.grade_class:
                    combined_grade_class = f"{record.grade.grade_number}{record.grade_class.name}"
                    teacher_data[teacher.id]["grades_classes"].add(combined_grade_class)

        # Transform sets into comma-separated strings for output
        results = []
        for teacher_id, data in teacher_data.items():
            results.append({
                "teacher_id": teacher_id,
                "teacher_name": data["teacher_name"],
                "email": data["email"],
                "mobile": data["mobile"],
                "school_name": data["school_name"],
                "subjects": ", ".join(sorted(data["subjects"])),
                "grades_classes": ", ".join(sorted(data["grades_classes"])),  # Updated field
            })
        download = request.query_params.get("download", "false").lower() == "true"
        if download:
            output = io.BytesIO()
            wb = Workbook()
            ws = wb.active
            ws.title = "Teachers Report"

            # Define headers for the Excel file
            headers = ["Teacher ID", "Teacher Name", "Email", "Mobile", "School Name", "Subjects", "Grades & Classes"]
            ws.append(headers)

            # Append data rows
            for row in results:
                ws.append([
                    row["teacher_id"],
                    row["teacher_name"],
                    row["email"],
                    row["mobile"],
                    row["school_name"],
                    row["subjects"],
                    row["grades_classes"],  # This will now show "10A, 12B"
                ])

            # Save the workbook and return as a file response
            wb.save(output)
            output.seek(0)
            return FileResponse(output, as_attachment=True, filename="teachers_report.xlsx")

        # Paginate the results
        paginator = self.pagination_class()
        paginated_data = paginator.paginate_queryset(results, request)
        return paginator.get_paginated_response(paginated_data)

    @action(detail=False, methods=["get"], url_path="school-learners-report")
    @permission_required('reports_management', 'view')
    def export_learners_with_parents(self, request, *args, **kwargs):
        # Determine the `school_id` based on user role
        if request.user.role.role_name in settings.SUPER_ROLE:
            school_id = request.query_params.get("school_id")
            if not school_id:
                return Response({"error": _("School ID is required for SUPER_ROLE users.")}, status=400)
        else:
            school_id = request.user.school_id

        # Validate batch_id
        batch_id = request.query_params.get("batch_id")
        if not batch_id:
            return Response({"error": _("Batch ID is required.")}, status=400)
        try:
            batch_id = int(batch_id)
        except ValueError:
            return Response({"error": _("Invalid batch_id format. It must be an integer.")}, status=400)

        # Validate school
        try:
            school = School.objects.get(id=school_id)
        except School.DoesNotExist:
            return Response({"error": _("School not found.")}, status=404)

        # Query `ClassSubjects`
        queryset = ClassSubject.objects.filter(
            school=school,
            batch_id=batch_id
        ).select_related("grade", "grade_class", "subject")

        # Apply search functionality for students
        search_query = request.query_params.get("search")
        if search_query:
            filtered_students = Student.objects.filter(
                Q(first_name__icontains=search_query) |
                Q(last_name__icontains=search_query) |
                Q(email__icontains=search_query) |
                Q(addmission_no__icontains=search_query)
            )
            queryset = queryset.prefetch_related(
                Prefetch("students", queryset=filtered_students, to_attr="filtered_students")
            )
        else:
            queryset = queryset.prefetch_related("students__parents")

        # Apply filters
        grade = request.query_params.get("grade")
        if grade:
            queryset = queryset.filter(grade=grade)

        grade_class = request.query_params.get("grade_class")
        if grade_class:
            queryset = queryset.filter(grade_class=grade_class)

        subject_id = request.query_params.get("subject_id")
        if subject_id:
            try:
                subject_id = int(subject_id)
                queryset = queryset.filter(subject=subject_id)
            except ValueError:
                return Response({"error": _("Invalid subject_id format. It must be an integer.")}, status=400)

        # Apply ordering
        order_by = request.query_params.get("ordering", "students__first_name")  # Default ordering by learner name
        reverse = order_by.startswith("-")
        if reverse:
            order_by = order_by[1:]
        valid_order_fields = ["students__first_name", "students__last_name", "students__email", "students__id", "students__addmission_no"]
        if order_by in valid_order_fields:
            queryset = queryset.order_by("-" + order_by if reverse else order_by)

        # Data aggregation for learners and parents
        learners_data = {}
        for cs in queryset:
            students_list = cs.filtered_students if search_query else cs.students.all()
            for learner in students_list:
                if learner.id not in learners_data:
                    learners_data[learner.id] = {
                        "learner_id": learner.id,
                        "admission_no": learner.addmission_no,
                        "learner_name": f"{learner.first_name} {learner.last_name}",
                        "learner_email": learner.email,
                        "learner_mobile": learner.mobile_number,
                        "grades_classes": set(),  # Combine grade and class here
                        "subjects": set(),
                        "parents": []
                    }

                # Combine grades and classes as `10A, 12B`
                if cs.grade and cs.grade_class:
                    combined_grade_class = f"{cs.grade.grade_number}{cs.grade_class.name}"
                    learners_data[learner.id]["grades_classes"].add(combined_grade_class)

                # Add subjects
                if cs.subject and cs.subject.master_subject:
                    subject_detail = (
                        f"{cs.subject.master_subject.subject_name} ({cs.subject.master_subject.subject_code})"
                    )
                    learners_data[learner.id]["subjects"].add(subject_detail)

                # Add parent information
                for parent in learner.parents.all():
                    parent_data = {
                        "parent_id": parent.id,
                        "parent_name": f"{parent.first_name} {parent.last_name}",
                        "parent_email": parent.email,
                        "parent_mobile": parent.mobile_number,
                    }
                    if parent_data not in learners_data[learner.id]["parents"]:
                        learners_data[learner.id]["parents"].append(parent_data)

        # Transform sets into lists and prepare the final response
        results = [
            {
                "learner_id": learner_id,
                "admission_no": data["admission_no"],
                "learner_name": data["learner_name"],
                "learner_email": data["learner_email"],
                "learner_mobile": data["learner_mobile"],
                "grades_classes": ", ".join(sorted(data["grades_classes"])),  # Updated to `10A, 12B`
                "subjects": ", ".join(sorted(data["subjects"])),
                "parents": data["parents"]
            }
            for learner_id, data in learners_data.items()
        ]

        # Download as Excel file
        download = request.query_params.get("download", "false").lower() == "true"
        if download:
            output = io.BytesIO()
            wb = Workbook()
            ws = wb.active
            ws.title = "Learners Report"

            # Headers for the Excel file
            headers = ["Learner ID", "Admission No", "Learner Name", "Email", "Mobile", "Grades & Classes", "Subjects", "Parents"]
            ws.append(headers)

            for row in results:
                parent_details = "; ".join(
                    [f"{p['parent_name']} ({p['parent_email']}, {p['parent_mobile']})" for p in row["parents"]]
                )
                ws.append([
                    row["learner_id"],
                    row["admission_no"],
                    row["learner_name"],
                    row["learner_email"],
                    row["learner_mobile"],
                    row["grades_classes"],  # Now showing `10A, 12B`
                    row["subjects"],
                    parent_details,
                ])

            wb.save(output)
            output.seek(0)
            return FileResponse(output, as_attachment=True, filename="learners_report.xlsx")

        # Paginate results
        paginator = self.pagination_class()
        paginated_data = paginator.paginate_queryset(results, request)
        return paginator.get_paginated_response(paginated_data)

    @action(detail=False, methods=["get"], url_path="school-parents-report")
    @permission_required('reports_management', 'view')
    def export_parents_with_learners(self, request, *args, **kwargs):
        # Determine the `school_id` based on the user's role
        if request.user.role.role_name in settings.SUPER_ROLE:
            school_id = request.query_params.get("school_id")
            if not school_id:
                return Response({"error": _("School ID is required for SUPER_ROLE users.")}, status=400)
        else:
            school_id = request.user.school_id

        # Get the year from the request (default to the current year)
        batch_id = request.query_params.get("batch_id")
        if not batch_id:
            return Response({"error": _("Batch ID is required.")}, status=400)
        try:
            batch_id = int(batch_id)
        except ValueError:
            return Response({"error": _("Invalid batch_id format. It must be an integer.")}, status=400)

        # Validate school
        if not school_id:
            return Response({"error": _("School ID is required.")}, status=400)

        try:
            school = School.objects.get(id=school_id)
        except School.DoesNotExist:
            return Response({"error": _("School not found.")}, status=404)

        # Query `ClassSubjects` and filter by year, school, and relationships
        queryset = ClassSubject.objects.filter(
            school=school,
            batch_id=batch_id
        ).select_related("grade", "grade_class", "subject").prefetch_related("students__parents")

        # Additional filters based on query parameters
        grade = request.query_params.get("grade")
        if grade:
            queryset = queryset.filter(grade=grade)

        grade_class = request.query_params.get("grade_class")
        if grade_class:
            queryset = queryset.filter(grade_class=grade_class)

        subject_id = request.query_params.get("subject_id")
        if subject_id:
            try:
                subject_id = int(subject_id)
                queryset = queryset.filter(subject=subject_id)
            except ValueError:
                return Response({"error": _("Invalid subject_id format. It must be an integer.")}, status=400)

        # Apply search filtering for parents by their name, email, or mobile
        search_query = request.query_params.get("search")
        if search_query:
            queryset = queryset.filter(
                Q(students__parents__first_name__icontains=search_query) |
                Q(students__parents__last_name__icontains=search_query) |
                Q(students__parents__email__icontains=search_query) |
                Q(students__parents__mobile_number__icontains=search_query)
            )

        # Data aggregation for parents and their learners
        parents_data = {}

        for cs in queryset:
            for learner in cs.students.all():
                for parent in learner.parents.all():
                    if parent.id not in parents_data:
                        parents_data[parent.id] = {
                            "parent_id": parent.id,
                            "parent_name": f"{parent.first_name} {parent.last_name}",
                            "parent_email": parent.email,
                            "parent_mobile": parent.mobile_number,
                            "learners": {},  # Store unique learners by ID
                        }

                    learner_id = learner.id
                    # Add learner details if not already added
                    if learner_id not in parents_data[parent.id]["learners"]:
                        parents_data[parent.id]["learners"][learner_id] = {
                            "learner_id": learner_id,
                            "admission_no": learner.addmission_no,
                            "name": f"{learner.first_name} {learner.last_name}",
                            "subjects": set(),
                            "grades_classes": set(),  # Combine grade and class
                        }

                    # Combine grades and classes as `10A, 12B`
                    if cs.grade and cs.grade_class:
                        combined_grade_class = f"{cs.grade.grade_number}{cs.grade_class.name}"
                        parents_data[parent.id]["learners"][learner_id]["grades_classes"].add(combined_grade_class)

                    # Add subject information to the learner
                    if cs.subject and cs.subject.master_subject:
                        subject_detail = (
                            f"{cs.subject.master_subject.subject_name} ({cs.subject.master_subject.subject_code})"
                        )
                        parents_data[parent.id]["learners"][learner_id]["subjects"].add(subject_detail)

        # Transform sets into lists and organize the final response
        results = []
        for parent_id, data in parents_data.items():
            learners = []
            for learner_data in data["learners"].values():
                learners.append({
                    "learner_id": learner_data["learner_id"],
                    "admission_no": learner_data["admission_no"],
                    "name": learner_data["name"],
                    "grades_classes": ", ".join(sorted(learner_data["grades_classes"])),  # Updated format
                    "subjects": ", ".join(sorted(learner_data["subjects"])),
                })

            results.append({
                "parent_id": parent_id,
                "parent_name": data["parent_name"],
                "parent_email": data["parent_email"],
                "parent_mobile": data["parent_mobile"],
                "learners": learners  # List of learners associated with the parent
            })
        download = request.query_params.get("download", "false").lower() == "true"
        if download:
            output = io.BytesIO()
            wb = Workbook()
            ws = wb.active
            ws.title = "Parents Report"

            # Create headers for the Excel file
            headers = ["Parent ID", "Parent Name", "Parent Email", "Parent Mobile", "Learners"]
            ws.append(headers)

            # Write rows to the Excel sheet
            for row in results:
                learners_data = []
                for learner in row["learners"]:
                    learners_data.append(
                        f'{learner["name"]} (Admission No: {learner["admission_no"]}, Grades & Classes: {learner["grades_classes"]}, Subjects: {learner["subjects"]})'
                    )

                # Create a single string for learners
                learners_str = "\n".join(learners_data)

                # Append the row to the worksheet
                ws.append([
                    row["parent_id"],
                    row["parent_name"],
                    row["parent_email"],
                    row["parent_mobile"],
                    learners_str
                ])

            wb.save(output)
            output.seek(0)
            return FileResponse(output, as_attachment=True, filename="parents_report.xlsx")
        paginator = self.pagination_class()
        paginated_data = paginator.paginate_queryset(results, request)
        return paginator.get_paginated_response(paginated_data)

    @action(detail=False, methods=["get"], url_path="school-assessments-report")
    @permission_required('reports_management', 'view')
    def export_school_assessments(self, request, *args, **kwargs):
        if request.user.role.role_name in settings.SUPER_ROLE:
            school_id = request.query_params.get("school_id")
        else:
            school_id = request.user.school_id

        # Get the year, defaulting to the current year
        batch_id = request.query_params.get("batch_id")
        if not batch_id:
            return Response({"error": _("Batch ID is required.")}, status=400)

        if not school_id:
            return Response({"error": _("School ID is required.")}, status=400)

        try:
            # Fetch all assessments for the given school and year
            assessments = Assessment.objects.filter(
                school_id=school_id,
                batch_id=batch_id
            ).prefetch_related('grade_class', 'grade', 'batch', 'term')

            # Apply filters from query parameters
            grade = request.query_params.get("grade")
            if grade:
                assessments = assessments.filter(grade=grade)

            grade_class = request.query_params.get("grade_class")
            if grade_class:
                assessments = assessments.filter(grade_class=grade_class)

            subject_id = request.query_params.get("subject_id")
            if subject_id:
                try:
                    subject_id = int(subject_id)
                    assessments = assessments.filter(
                        id__in=AssessmentSubject.objects.filter(
                            subject_id=subject_id
                        ).values_list('assessment_id', flat=True)
                    )
                except ValueError:
                    return Response({"error": _("Invalid subject_id format. It must be an integer.")}, status=400)

            # Structure the JSON response
            data = []
            for assessment in assessments:
                # Assessment details
                assessment_status = assessment.status
                assessment_name = assessment.assessment_name

                # Fetch related grade classes and format as `10A, 12B`
                grade_classes = [
                    f"{gc.grade.grade_number}{gc.name}" for gc in assessment.grade_class.all()
                ]
                grade_classes_str = ", ".join(sorted(grade_classes))  # Convert to `10A, 12B`

                # Get **distinct** subject names to prevent duplicates
                subject_names = list(
                    AssessmentSubject.objects.filter(assessment_id=assessment.id)
                    .values_list("subject__master_subject__subject_name", flat=True)
                    .distinct()
                )
                subjects_str = ", ".join(subject_names)  # ✅ Convert list to string for Excel

                # Total students and attendance data
                total_students_qs = Student.objects.filter(
                    grade_class__assessments=assessment
                ).distinct()
                total_students = total_students_qs.count()

                students_attended_qs = StudentAnswerSheet.objects.filter(
                    assessment_subject__assessment=assessment
                ).values_list("student_id", flat=True).distinct()
                attended_count = len(students_attended_qs)
                not_attended_count = total_students - attended_count

                # Aggregate marks data
                marks_data = StudentAnswerSheet.objects.filter(
                    assessment_subject__assessment=assessment
                ).aggregate(
                    total_marks=Sum("obtained_mark"),
                    average_marks=Avg("obtained_mark")
                )

                total_marks = marks_data["total_marks"] or 0
                average_marks = marks_data["average_marks"] or 0

                # Append the assessment data to the response list
                data.append({
                    "assessment_name": assessment_name,
                    "subjects": subject_names,  # ✅ Keep list for JSON API response
                    "subjects_excel": subjects_str,  # ✅ Convert to string for Excel
                    "grade_classes": grade_classes_str,  # ✅ Now formatted as `10A, 12B`
                    "status": assessment_status,
                    "total_students": total_students,
                    "students_attended_all": attended_count,
                    "students_not_attended_all": not_attended_count,
                    "total_marks": total_marks,
                    "average_marks": round(average_marks, 2),
                })
            download = request.query_params.get("download", "false").lower() == "true"
            if download:
                output = io.BytesIO()
                wb = Workbook()
                ws = wb.active
                ws.title = "Assessments Report"

                # Create headers from the keys in the first data item
                headers = ["Assessment Name", "Subjects", "Grade Classes", "Status",
                           "Total Students", "Students Attended", "Students Not Attended", "Total Marks",
                           "Average Marks"]
                ws.append(headers)

                # Write rows to the worksheet
                for row in data:
                    ws.append([
                        row["assessment_name"],
                        row["subjects_excel"],  # ✅ Use string format for Excel
                        row["grade_classes"],  # ✅ Now shows `10A, 12B`
                        row["status"],
                        row["total_students"],
                        row["students_attended_all"],
                        row["students_not_attended_all"],
                        row["total_marks"],
                        row["average_marks"]
                    ])

                wb.save(output)
                output.seek(0)
                return FileResponse(output, as_attachment=True, filename="assessments_report.xlsx")

            # Sort and paginate response
            paginator = self.pagination_class()
            paginated_data = paginator.paginate_queryset(data, request)
            return paginator.get_paginated_response(paginated_data)

        except School.DoesNotExist:
            return Response({"error": _("School not found.")}, status=404)

        except Exception as e:
            return Response(
                {"error": _("An error occurred: %(error)s") % {"error": str(e)}},
                status=500
            )

    @action(detail=False, methods=["get"], url_path="school-quizzes-report")
    @permission_required('reports_management', 'view')
    def export_quiz_report(self, request, *args, **kwargs):
        if request.user.role.role_name in settings.SUPER_ROLE:
            school_id = request.query_params.get("school_id")
        else:
            school_id = request.user.school_id

        # Default to the current year
        batch_id = request.query_params.get("batch_id")
        if not batch_id:
            return Response({"error": _("Batch ID is required.")}, status=400)

        if not school_id:
            return Response({"error": _("School ID is required.")}, status=400)

        try:
            # Fetch all `MainQuiz` instances filtered by school and year
            main_quizzes = MainQuiz.objects.filter(
                school_id=school_id,
                batch_id=batch_id
            ).select_related('subject', 'grade', 'grade_class', 'term', 'batch', 'school')

            # Apply filters from query params
            grade = request.query_params.get("grade")
            if grade:
                main_quizzes = main_quizzes.filter(grade=grade)

            grade_class = request.query_params.get("grade_class")
            if grade_class:
                main_quizzes = main_quizzes.filter(grade_class=grade_class)

            subject_id = request.query_params.get("subject_id")
            if subject_id:
                try:
                    subject_id = int(subject_id)
                    main_quizzes = main_quizzes.filter(subject_id=subject_id)
                except ValueError:
                    return Response({"error": _("Invalid subject_id format. It must be an integer.")}, status=400)

            # JSON response data structure
            data = []

            # Loop through all fetched `MainQuiz` instances to process data
            for main_quiz in main_quizzes:
                # Combine subject name and subject's code
                subject_name = main_quiz.subject.master_subject.subject_name
                subject_code = main_quiz.subject.master_subject.subject_code
                subject_with_code = f"{subject_name} ({subject_code})"

                # Get all associated `Quiz` instances and calculations
                assigned_quizzes = Quiz.objects.filter(main_quiz=main_quiz)
                assigned_students = assigned_quizzes.count()  # Total assigned quizzes / students
                completed_students = assigned_quizzes.filter(is_attempted=True).count()  # Completed quizzes
                not_completed_students = assigned_students - completed_students  # Students who haven't completed

                # Calculate available and average marks
                total_marks_per_quiz = main_quiz.number_of_questions  # Assuming each question is 1 mark
                avg_marks = assigned_quizzes.aggregate(average_marks=Avg('marks_obtained'))["average_marks"] or 0

                # Combine grade and grade_class as `10A, 12B`
                grade_class_combined = f"{main_quiz.grade.grade_number}{main_quiz.grade_class.name}"

                # Append the data to the JSON response
                data.append({
                    "id": main_quiz.id,
                    "quiz_title": main_quiz.title,
                    "grade_class": grade_class_combined,  # Updated format
                    "subject": subject_with_code,
                    "total_questions": main_quiz.number_of_questions,
                    "students_assigned": assigned_students,
                    "students_completed": completed_students,
                    "students_not_completed": not_completed_students,
                    "total_marks_available": total_marks_per_quiz,
                    "average_marks_obtained": round(avg_marks, 2),
                })
            download = request.query_params.get("download", "false").lower() == "true"
            if download:
                output = io.BytesIO()
                wb = Workbook()
                ws = wb.active
                ws.title = "Quizzes Report"

                # Create Excel headers
                headers = ["Quiz ID", "Quiz Title", "Grade & Class", "Subject", "Total Questions",
                           "Students Assigned", "Students Completed", "Students Not Completed",
                           "Total Marks Available", "Average Marks Obtained"]
                ws.append(headers)

                # Add rows to the worksheet
                for row in data:
                    ws.append([
                        row["id"],
                        row["quiz_title"],
                        row["grade_classes"],  # Now showing `10A, 12B`
                        row["subject"],
                        row["total_questions"],
                        row["students_assigned"],
                        row["students_completed"],
                        row["students_not_completed"],
                        row["total_marks_available"],
                        row["average_marks_obtained"]
                    ])

                # Save the file and return a response
                wb.save(output)
                output.seek(0)
                return FileResponse(output, as_attachment=True, filename="quizzes_report.xlsx")
            # Paginate the JSON response
            paginator = self.pagination_class()
            paginated_data = paginator.paginate_queryset(data, request)
            return paginator.get_paginated_response(paginated_data)

        except School.DoesNotExist:
            return Response({"error": _("School not found.")}, status=404)

        except Exception as e:
            return Response(
                {"error": _("An error occurred: %(error)s") % {"error": str(e)}},
                status=500
            )
